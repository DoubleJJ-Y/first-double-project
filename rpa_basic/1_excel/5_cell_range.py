from openpyxl import Workbook
from random import *

wb = Workbook()
ws =  wb.active

# 1줄씩 데이터 넣기 (리스트나 튜플로 넣음)
ws.append(["번호", "영어", "수학"])
for i in range(1,11): #10개 데이터 넣기
    ws.append(([i, randint(0,100), randint(0,100) ]))

# #column value 가져오기
# col_B = ws["B"] #영어 column 만 가지고 오기
# for cell in col_B:
#     print(cell.value)

# col_range = ws["A:C"] #A~C까지의 Data가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# #row value 가져오기
# row_title = ws[1] #Row는 숫자. 문자는 열.
# print(row_title)

# row_range = ws[2:6] #2번째 줄에서 6번째 줄까지 가지고 오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end= " ")
#     print()

# #row의 끝을 모를 때
# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]
# for rows in row_range:    
#     for cell in rows:
#         print(cell.coordinate, end=" ") #좌표를 가져올 수 있다.
#         xy = coordinate_from_string(cell.coordinate)
#         #print(xy, end=" ")
#         #print(xy, end=" ")
#         #print(xy, end=" ")
#     print()
# for rows in row_range:    
#     for cell in rows:
#         print(cell.coordinate, end=" ") #좌표를 가져올 수 있다.
#         xy = coordinate_from_string(cell.coordinate)
#         #print(xy, end=" ")
#         print(xy[0], end=" ") # A
#         print(xy[1], end=" ") # 2 행 열 정보 각각 알아오게하기 
#     print()


# ##튜플??
# # 리스트는 [], 튜플은 ()으로 둘러싼다.
# # 리스트는 요솟값의 생성, 삭제, 수정이 가능하지만, 튜플은 요솟값을 바꿀 수 없다.

# #한 row 씩 가져와서 tuple로.
# print(tuple(ws.rows))

# #한 columns 씩 가져와서 tuple로.
# print(tuple(ws.columns))

# #tuple로 값 print하기
# #row
# for row in tuple(ws.rows):
#     print(row[1].value)
# #column
# for row in tuple(ws.coluumns):
#     print(row[1].value)

# #iterator(반복자) 사용하기
# for row in ws.iter_rows(): #전체 row
#     print(row[1].value)

# for column in ws.iter_cols(): #전체 column
#     print(column[0].value)


#iterator(반복자) 사용하기 - 슬라이싱 
# #행 좌,우 로 Data를 받음 (<Cell 'Sheet'.B2>, <Cell 'Sheet'.C2>)
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): #슬라이싱 느낌 가능
#     print(row[0].value,row[1].value) #수학, 영어
#     #print(row)

# #열 상,하 로 Data를 받음 (<Cell 'Sheet'.B2>, <Cell 'Sheet'.B3>,...
# for col in ws.iter_cols(min_row=2, max_row=11, min_col=2, max_col=3): #슬라이싱 느낌 가능
#     #print(row[0].value,row[1].value) #수학, 영어
#     print(col)

wb.save("sample.xlsx")
