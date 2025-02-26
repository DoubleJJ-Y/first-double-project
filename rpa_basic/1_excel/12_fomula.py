import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active 

#엑셀과 똑같은 수식으로 셀을 설정할 수 있다.
ws["A1"] = datetime.datetime.today() #오늘 날짜 정보
ws["A2"] = "=SUM(1,2,3)" # 1+2+3 = 6
ws["A3"] = "=AVERAGE(1,2,3,)" 

ws["A4"] =  10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

wb.save("sample_formualr.xlsx")
