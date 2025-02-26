# Quiz) 여러분은 나도대학의 컴퓨터과 교수님입니다.
# 여러분이 가르치는 과목의 점수 비중은 다음과 같습니다.

# - 출석 10
# - 퀴즈1 10
# - 퀴즈2 10
# - 중간고사 20
# - 기말고사 30
# - 프로젝트 20
# ------------
# - 총합 100

# 마지막 수업을 모두 마치고 이번 학기 학생들의 최종 성적을 검토하는 과정에서
# 퀴즈2 문제에 오류를 발견하여 모두 만점 처리를 하기로 하였습니다.
# 현재까지 작성된 최종 성적 데이터를 기준으로 아래와 같이 수정하시오.

# 1. 퀴즈2 점수를 10 으로 수정
# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
# - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
# 3. 출석이 5 미만인 학생은 총점 상관없이 F

# ※ 최종 파일명 : scores.xlsx

# [현재까지 작성된 최종 성적 데이터]
# 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
# 1,10,8,5,14,26,12
# 2,7,3,7,15,24,18
# 3,9,5,8,8,12,4
# 4,7,8,7,17,21,18
# 5,7,8,7,16,25,15
# 6,3,5,8,8,17,0
# 7,4,9,10,16,27,18
# 8,6,6,6,15,19,17
# 9,10,10,9,19,30,19
# 10,9,8,8,20,25,20

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook("score_org.xlsx")
ws = wb.active

#ws.append(["학번", "츨석", "퀴즈", "퀴즈2", "중간고사", "기말고사", "프로젝트", "총점","등급"]) 이렇게 하면 줄로 추가가 된다!
ws["H1"].value = "총점"
ws["I1"].value = "등급"


#총합 열 구하기
#H2 =SUM(B2:G2)
ws["H2"] =  "=SUM(B2:G2)"
ws["H3"] =  "=SUM(B3:G3)"
ws["H4"] =  "=SUM(B4:G4)"
ws["H5"] =  "=SUM(B5:G5)"
ws["H6"] =  "=SUM(B6:G6)"
ws["H7"] =  "=SUM(B7:G7)"
ws["H8"] =  "=SUM(B8:G8)"
ws["H9"] =  "=SUM(B9:G9)"
ws["H10"] = "=SUM(B10:G10)"
ws["H11"] = "=SUM(B11:G11)"

#퀴즈2는 모두 만점
ws["D2"] =  10
ws["D3"] =  10
ws["D4"] =  10
ws["D5"] =  10
ws["D6"] =  10
ws["D7"] =  10
ws["D8"] =  10
ws["D9"] =  10
ws["D10"] = 10
ws["D11"] = 10

#=IF(H2>90,"A",IF(H2>80,"B",IF(H2>70,"C","D")))
# ws["I2"] = "=IF(H2>90,"A",IF(H2>80,"B",IF(H2>70,"C","D")))"
# ws["I3"] = "=IF(H3>90,"A",IF(H3>80,"B",IF(H3>70,"C","D")))"
# ws["I4"] = "=IF(H4>90,"A",IF(H4>80,"B",IF(H4>70,"C","D")))"
# ws["I5"] = "=IF(H5>90,"A",IF(H5>80,"B",IF(H5>70,"C","D")))"
# ws["I6"] = "=IF(H6>90,"A",IF(H6>80,"B",IF(H6>70,"C","D")))"
# ws["I7"] = "=IF(H7>90,"A",IF(H7>80,"B",IF(H7>70,"C","D")))"
# ws["I8"] = "=IF(H8>90,"A",IF(H8>80,"B",IF(H8>70,"C","D")))"
# ws["I9"] = "=IF(H9>90,"A",IF(H9>80,"B",IF(H9>70,"C","D")))"
# ws["I10"]= "=IF(H10>90,"A",IF(H10>80,"B",IF(H10>70,"C","D")))"
# ws["I11"]= "=IF(H11>90,"A",IF(H11>80,"B",IF(H11>70,"C","D")))"

# for row in ws.iter_rows(min_row=2):
#     #번호, 영어, 수학
#     if int(row[1].value) >80 :
#         print(row[0].value, "번 학생은 영어 천재")

# #
# for row in ws.iter_rows(max_row=1): #
#     for cell in row:
#         if cell.value =="영어":
#             cell.value == "컴퓨터"

wb.save("score.xlsx")