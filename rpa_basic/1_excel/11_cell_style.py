from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

#번호, 영어, 수학
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]


#셀의 크기
ws.column_dimensions["A"].width = 5 #열은 넓이
ws.row_dimensions[1].height = 50 #행은 높이


#스타일 적용
a1.font = Font(color="FF0000", italic = True, bold = True)
b1.font = Font(color="00FF00", name= "Arial", strike= True) #
c1.font = Font(color="0000FF", size = 20, underline= "single")

#테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin") )
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 70 이상 셀에 대해서 초록색 적용
#전체셀
for row in ws.rows :
    for cell in row :
        cell.alignment = Alignment(horizontal="center", vertical="center") #위/아래, 좌/우 중앙 정렬
        #center, left, right, top, bottom
        if cell.column == 1: # A번호열은 제외
            continue
        if isinstance(cell.value, int) and cell.value > 80 : #정수형이고 80보다 크면
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") #배경을 초록색
            cell.font = Font(color="FF0000") # 폰트 색 변경

# 틀 고정
ws.freeze_panes = "B2" #B2 기준 틀고정

wb.save("sample_style.xlsx")