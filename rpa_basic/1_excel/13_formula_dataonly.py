from openpyxl import load_workbook

# #수식 그대로 가져오게 됨.
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# for row in ws.valuse:
#     for cell in row:
#         print(cell)


# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 값을 가져옴
wb = load_workbook("sample_formula.xlsx", data_only=True) #이렇게 열어줘야 함.
ws = wb.active

#수식 그대로 가져오게 됨.
for row in ws.valuse:
    for cell in row:
        print(cell)
