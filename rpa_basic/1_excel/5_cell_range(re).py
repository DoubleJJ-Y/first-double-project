from openpyxl import Workbook
from random import *

wb = Workbook()
ws =  wb.active

# 1줄씩 데이터 넣기 (리스트나 튜플로 넣음)
ws.append(["번호", "영어", "수학"])
for i in range(1,11): #10개 데이터 넣기
    ws.append(([i, randint(0,100), randint(0,100) ]))

# col_B = ws["B"] #B 컬럼 모든 변수 다 저장.(*있는 값만 다 가져오네.)
# #print(col_B)

# for cell in col_B:
#     print(cell.value)


# #B, C 열 모든 값 가져오기
# col_range = ws["B:C"] #영어, 수학 column 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

##이번엔 행으로 가져와보자

# #1행 모든 값 가져오기
# row_title = ws[1]
# for cell in row_title:
#     print(cell.value)

# #1~6 행 값 가져오기
# row_range = ws[1:6]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print() # 줄바꿈

# #행의 끝까지 가져오기
# row_range = ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print() #줄바꿈

# #좌표 정보 가져오기 / 값만 가져오면 그게 어디에 있는 값인지 알 수 없다!
# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         print(cell.coordinate, end=" ") #A10, ZC250
#         xy = coordinate_from_string(cell.coordinate) #xy좌표를 각 xy[0],xy[1]이 가지고 있어서 나중에 쓰임새가 좋다.
#         print(xy[0], end=" ")
#         print(xy[1], end=" ")
#         #print(xy, end = " ")
#     print() #줄바꿈


### Tuple 형
# # 전체 rows 가져오기
# print(tuple(ws.rows)) #한 줄씩 가져와서 tuple #A1 A2 ~

# # 전체 rows 가져오기
# print(tuple(ws.columns)) #한 열씩 가져와서 tuple  #A1 B2 ~

# for row in tuple(ws.rows): #A1, A2 ~ 이렇게 가져오니까 지정한 열을 가져온다. (*역생각)
#     print(row[1].value) #0 :번호 열, 1:영어 열, 2.수학 열

# for column in tuple(ws.columns): #A1, B2 ~ 이렇게 가져오니까 지정한 행을 가져온다. (*역생각)
#     print(column[0].value) # 0:머릿글행 번호 영어 수학

##iter(반복자) 를 이용하면 tuple(ws.columns 위와 동일한 결과 / 하지만 변수를 넣어줄 수 있어서 활용성 Up!
# for column in ws.iter_cols():
#     print(column[0].value)



# for row in ws.iter_rows():
#     print(row[0].value)

# 2~11번 째 줄까지, 2~3번째 열까지
# test data csv 처럼 47번쨰 줄부터 사용한다 하면 min_row = 47
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    #print(row[0].value, row[1].value)
    print(row)



wb.save("sample.xlsx")
